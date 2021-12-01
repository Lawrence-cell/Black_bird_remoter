# -*- coding: utf-8-*-

import openpyxl
import os
from math import log10 as log

# 导出自Energy History (.his)
# 读取text文件
f_th = open('threshold.thd', 'r')

text_th = f_th.readlines()
len_th = len(text_th)

f_th.close()


# 保存文件
wk_before = openpyxl.Workbook()
wt_th = wk_before.active
wt_th.title = 'threshold'


title = ['Frequency/MHz',' Maximum Amplitude/dBm', 'Threshold Level/dBm']


for i in range(1,4):
    wt_th.cell(row=1, column=i, value=title[i-1])


for i in range(2, len_th):
    mid_str = text_th[i].split(",")
    k = len(mid_str)
    # print(k)
    # print(mid_str)


    for j in range(k-1):

        h = ((k-1)*(i-2) + (j+1))

        frequency = 20000000 + 6835.94 * (h-1)
        wt_th.cell(h+1, 1, frequency/1E6)

        value_Vpp = float(mid_str[j])
        value_dBm = 10*log(value_Vpp * 0.5)-10*log(50)+30
        wt_th.cell(h+1, 2, value_dBm)
        wt_th.cell(h+1, 3, -96.0)

if os.path.exists("./Excel_Threshold.xlsx"):
    os.remove("./Excel_Threshold.xlsx")
    wk_before.save('Excel_Threshold.xlsx')
else:
    wk_before.save('Excel_Threshold.xlsx')

